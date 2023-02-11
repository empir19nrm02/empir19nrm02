########################################################################
# <MCLoadSave: a Python module for MC simulations.>
# Copyright (C) <2023>  <Udo Krueger> (udo.krueger at technoteam.de)
#########################################################################

"""
Module to read and save similation results
===========================================================

.. codeauthor:: UK
"""

from empir19nrm02.MC import  MCVectorVar
import numpy as np
import pandas as pd
import os
from numpy import ndarray


__all__ = ['load_from_excel','save_to_excel','save_to_csv', 'load_from_csv', 'load_from_excel_raw']

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_data_from_array(allCells:ndarray, wlColumn = 0, valColumn=1, stdColumn = None, covColumn=2, corrColumn = None, startRow=0):
    rownumber = allCells.shape[0]
    wl = allCells[startRow:rownumber, wlColumn]
    signal = allCells[startRow:rownumber, valColumn]
    mcVectorVar = MCVectorVar(elements=wl.shape[0])
    if covColumn:
        cov_matrix = allCells[startRow:rownumber, covColumn:rownumber+covColumn-startRow]
        mcVectorVar.set_vector_param(v_mean=signal, cov=cov_matrix)
    else:
        if corrColumn and stdColumn:
            std = allCells[startRow:rownumber, stdColumn]
            corr_matrix = allCells[startRow:rownumber, corrColumn:rownumber + corrColumn - startRow]
            mcVectorVar.set_vector_param(v_mean=signal, v_std=std, corr=corr_matrix)
    return wl, mcVectorVar

def is_valid_number(_str):
    if _str is None:
        return False
    try:
        float(_str)
        return True
    except ValueError:
        return False
def load_from_excel_raw(filename:str, worksheet:str=None):
    """
    Reads from excel raw data .

    Args:
        :filename:
            | filename
        :worksheet:
            | None or name or number, optional
            | If 'None': take the first worksheet
            | if number: take the worksheet with the number (starting with 0)
            | If name: take the worksheet with the name.

    Returns:
        :returns:
            | all cells as a ndarray

    Note:
        Attention: no error management
    """
    if not ('.xls' in filename or '.xlsx' in filename):
        filename = filename + '.xlsx'

    wb = load_workbook(filename = filename, data_only=True)
    if worksheet is None:
        worksheet = wb.sheetnames[0]
    if isinstance(worksheet, int):
        worksheet = wb.sheetnames[worksheet]
    else:
        worksheet = worksheet
    ws = wb[worksheet]
    allCells = np.array([[cell.value if is_valid_number(cell.value) else np.nan for cell in row] for row in ws.iter_rows()],dtype=float)
    allCellsStr = np.array([[cell.value for cell in row] for row in ws.iter_rows()],dtype=str)

    return allCellsStr, allCells
def load_from_excel(filename:str, worksheet:str=None, wlColumn = 1, valColumn=2, stdColumn = None, covColumn=4, corrColumn = None, startRow=2):
    """
    Reads data from excel file to create an instance of MCVectorVar.

    Args:
        :filename:
            | filename
        :worksheet:
            | None or name or number, optional
            | If 'None': take the first worksheet
            | if number: take the worksheet with the number (starting with 0)
            | If name: take the worksheet with the name.
        :wlColumn:
            | Column number for the wavelength data.
        :valColumn:
            | Column number for the value / signal data.
        :stdColumn:
            | Column number for the standard deviation  data.
        :covColumn:
            | Column number for the start of the covariance matrix.
        :corrColumn:
            | Column number for the start of the covariance matrix.

    Returns:
        :returns:
            | wavelength vector
            | instance of MCVectorVar

    Note:
        The covMatrix or the stdDev Vector and the correlation matrix is needed
        Attention: no error management
    """

    _, allCells = load_from_excel_raw(filename, worksheet=worksheet)

    # sometimes there are some nan lines appended at the end of the array
    if np.isnan(allCells[startRow:,wlColumn]).any():
        rownumber = np.where(np.isnan(allCells[startRow:,wlColumn]))[0][0]
    else:
        rownumber = allCells.shape[0]

    return extract_data_from_array(allCells[0:rownumber,:], wlColumn=wlColumn, valColumn=valColumn, stdColumn=stdColumn, covColumn=covColumn, corrColumn=corrColumn, startRow=startRow)


def load_from_csv( filename:str, wlColumn = 1, valColumn=2, stdColumn = None, covColumn=4, corcColumn = None, startRow=1, delimiter=';'):
    """
    Reads data from csv file to create an instance of MCVectorVar.

    Args:
        :filename:
            | filename
        :wlColumn:
            | Column number for the wavelength data.
        :valColumn:
            | Column number for the value / signal data.
        :stdColumn:
            | Column number for the standard deviation  data.
        :covColumn:
            | Column number for the start of the covariance matrix.
        :corrColumn:
            | Column number for the start of the covariance matrix.
        :delimiter:
            | delimiter between the numbers

    Returns:
        :returns:
            | wavelength vector
            | instance of MCVectorVar

    Note:
        The covMatrix or the stdDev Vector and the correlation matrix is needed
        Automatic handling of point/comma
        Attention: no error management
    """



    if not '.csv' in filename:
        filename = filename + '.csv'

    def conv(x):
        return x.replace(',', '.').encode()

    allCells = np.genfromtxt((conv(x) for x in open(filename)), delimiter=delimiter)

    return extract_data_from_array(allCells, wlColumn=wlColumn, valColumn=valColumn, stdColumn=stdColumn, covColumn=covColumn, corrColumn=corcColumn, startRow=startRow)

def collect_data(wl:ndarray|MCVectorVar, mcvar:MCVectorVar, use_runData = True)->pd.DataFrame:
    """
    collect the data from the MCVectorVars to a pd.DataFrame.

    Args:
        :wl:
            | ndarray or MCVectorVar with wavelength information
        :mcVar:
            | McVectorVar with information to be saved (mean, stdDev, corr, cov information only)
        :use_runData:
            | Out the data calculated from the MC runs. Otherwise, the setData from the simulation

    Returns:
        :df:
            | pd.DataFrame with the collected information
    Note:
        Attention: no error management
    """

    # use the nominal wavelength information from the MCVectorVar only
    if isinstance( wl, MCVectorVar):
        wl = wl.val[0]

    # put all the data in a pd.DataFrame
    df_wl = pd.DataFrame(wl, columns = ['wl'])
    if use_runData:
        data = mcvar.runData
    else:
        data = mcvar.setData
    df_v_mean = pd.DataFrame(data.v_mean, columns = ['mean'])
    df_v_std = pd.DataFrame(data.v_std, columns = ['std'])
    cov_names = ['cov_' + str(i) for i in range(wl.shape[0])]
    df_cov = pd.DataFrame(data.cov_matrix, columns = cov_names)
    corr_names = ['corr_' + str(i) for i in range(wl.shape[0])]
    df_corr = pd.DataFrame(data.corr_matrix, columns = corr_names)

    df = pd.concat([df_wl, df_v_mean, df_v_std, df_cov, df_corr], axis=1)

    return df

def save_to_excel(wl:ndarray|MCVectorVar, mcvar:MCVectorVar, filename:str, worksheet:str=None, use_runData = True)->None:
    """
    save data to excel file from the  instance of MCVectorVar.

    Args:
        :wl:
            | ndarray or MCVectorVar with wavelength information
        :mcVar:
            | McVectorVar with information to be saved (mean, stdDev, corr, cov information only)
        :filename:
            | filename
        :worksheet:
            | None or name or number, optional
            | If 'None': take the first worksheet
            | if number: take the worksheet with the number (starting with 0)
            | If name: take the worksheet with the name.
        :use_runData:
            | Out the data calculated from the MC runs. Otherwise, the setData from the simulation

    Returns:
        :None:
    Note:
        The data will be stored in the worksheet starting with A1 without deleting
        other data if the worksheet exists
        Attention: no error management
    """

    if not ('.xls' in filename or '.xlsx' in filename):
        filename = filename + '.xlsx'

    if os.path.isfile( filename):
        wb = load_workbook(filename = filename, data_only=True)
    else:
        wb = Workbook()

    if worksheet:
        if isinstance(worksheet, str):
            if not worksheet in wb.sheetnames:
                wb.create_sheet(worksheet)
        else:
            worksheet = wb.sheetnames[worksheet]
    else:
        worksheet = wb.sheetnames[0]

    ws = wb[worksheet]

    df = collect_data(wl, mcvar, use_runData=use_runData)

    for i, r in enumerate( dataframe_to_rows(df, index=True, header=True), start=1):
        for j, text in enumerate(r, start=1):
            ws.cell(column=j, row=i, value=text)

    wb.save(filename)


def save_to_csv(wl:ndarray|MCVectorVar, mcvar:MCVectorVar, filename:str, sep=',', decimal='.', use_runData = True)->None:
    """
    save data to excel file from the  instance of MCVectorVar.

    Args:
        :wl:
            | ndarray or MCVectorVar with wavelength information
        :mcVar:
            | McVectorVar with information to be saved (mean, stdDev, corr, cov information only)
        :filename:
            | filename
        :sep:
            see pandas to_csv
        :delimiter:
            see pandas to_csv
        :use_runData:
            | Out the data calculated from the MC runs. Otherwise, the setData from the simulation

    Returns:
        :None:
    Note:
        The data will be stored in the worksheet starting with A1 without deleting
        other data if the worksheet exists
        Attention: no error management
    """

    df = collect_data(wl, mcvar, use_runData=use_runData)
    if not '.csv' in filename:
        filename = filename + '.csv'
    df.to_csv(filename, sep=sep, decimal=decimal)
